#!/usr/bin/env python3
"""
Export cost prices from "Price Adj September 2025.xlsx" for import into plain_products.
Sheet "Price Adj September 2025" has: Code, Product Group Description, Name, Current Price, New Price.
Current Price = cost price per case. Outputs one JSON object per line to stdout for scripts/import-cost-to-db.js.

Usage:
  pip install openpyxl   # if not already installed
  python scripts/export-cost-from-price-adj.py [path/to/Price Adj September 2025.xlsx]
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def main():
    repo_root = Path(__file__).resolve().parent.parent
    default_xlsx = repo_root / "Price Adj September 2025.xlsx"
    downloads = Path.home() / "Downloads" / "Price Adj September 2025.xlsx"
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else (default_xlsx if default_xlsx.exists() else downloads)
    if not xlsx_path.exists():
        print(f"Excel file not found: {xlsx_path}", file=sys.stderr)
        print("Usage: python scripts/export-cost-from-price-adj.py [path/to/Price Adj September 2025.xlsx]", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = "Price Adj September 2025"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        sys.exit(0)
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    code_col = next((i for i, h in enumerate(header) if "code" in h.lower() and "product" not in h.lower()), 0)
    current_price_col = next((i for i, h in enumerate(header) if "current price" in h.lower()), 3)

    count = 0
    for r in rows[1:]:
        if not r or len(r) <= max(code_col, current_price_col):
            continue
        code_raw = r[code_col]
        price_raw = r[current_price_col] if current_price_col < len(r) else None
        if code_raw is None or code_raw == "":
            continue
        code = str(code_raw).strip()
        if not code:
            continue
        try:
            if price_raw is None or price_raw == "":
                continue
            cost = float(price_raw)
        except (TypeError, ValueError):
            continue
        print(json.dumps({"code": code, "cost": round(cost, 2)}))
        count += 1

    print(f"# Exported {count} cost prices", file=sys.stderr)


if __name__ == "__main__":
    main()
